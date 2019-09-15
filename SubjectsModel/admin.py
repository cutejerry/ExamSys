from django.contrib import admin
from django.urls import path
from django import forms
from django.shortcuts import render, redirect
from django.contrib import messages
from SubjectsModel.models import Exam, Subject
import openpyxl

class exam_paper:
    def __init__(self, n):
        self.name = n
        self.subjects = []

    def add_sub(self, s):
        self.subjects.append(s)

class exam_subject:
    def __init__(self, Num, Pmt, A, B, C, D, E, Ans, Pic, Type, Period, Theme, P1, P2, P3, Logic, Difficulty):
        self.Number = Num
        self.Promt = Pmt
        self.A = A
        self.B = B
        self.C = C
        self.D = D
        self.E = E
        self.Answer = Ans
        self.Picture = Pic
        self.Type = Type
        self.Period = Period
        self.Theme = Theme
        self.Point1 = P1
        self.Point2 = P2
        self.Point3 = P3
        self.Logic = Logic
        self.Difficulty = Difficulty

#class SubjectInline(admin.StackedInline):
#    model = Subject

class XlsxImportForm(forms.Form):
    xlsx_file = forms.FileField()

class ExamAdmin(admin.ModelAdmin):
    change_list_template = 'exam_changelist.html'
    #inlines = [SubjectInline]

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-xlsx/', self.import_xlsx),
        ]
        return my_urls + urls

    def import_xlsx(self, request):
        if request.method == 'POST':
            xlsx_file = request.FILES['xlsx_file']
            if( self.process_xlsx(xlsx_file) != -1 ):
                self.message_user(request, "Your xlsx file has been imported successfully")
            else:
                self.message_user(request, "Your exam name has been exist, avoid importing", level=messages.WARNING)
            return redirect("..")
        form = XlsxImportForm()
        payload = {"form": form}
        return render(
            request, "xlsx_form.html", payload
        )

    def copy_sub(self, obj, sub):
        obj.number = sub.Number
        obj.content_txt  = sub.Promt
        obj.answer_idx = ord(sub.Answer) - 64
        obj.options1 = sub.A
        obj.options2 = sub.B
        obj.options3 = sub.C
        obj.options4 = sub.D
        obj.options5 = sub.E
        obj.type = sub.Type
        obj.period = sub.Period
        obj.theme = sub.Theme
        obj.point1 = sub.Point1
        obj.point2 = sub.Point2
        obj.point3 = sub.Point3
        obj.logic = sub.Logic
        obj.difficult = sub.Difficulty

    def process_xlsx(self, xlsx_file):
        print(self.model)
        print(type(xlsx_file))
        wb = openpyxl.load_workbook(xlsx_file)
        sheets = wb.sheetnames
        print(sheets[0])
        # for obj in objs:

        (exam_obj, created) = Exam.objects.get_or_create(name=sheets[0])
        if created != True:
            return -1
        else:
            ws = wb.active
            new_paper = exam_paper(sheets[0])

            for i in range(2, ws.max_row + 1):
                subject = exam_subject(Num=new_paper.name + '_' + ws['A' + str(i)].value, \
                                       Pmt=ws['B' + str(i)].value, \
                                       A=ws['C' + str(i)].value, \
                                       B=ws['D' + str(i)].value, \
                                       C=ws['E' + str(i)].value, \
                                       D=ws['F' + str(i)].value, \
                                       E=ws['G' + str(i)].value, \
                                       Ans=ws['H' + str(i)].value, \
                                       Pic=ws['I' + str(i)].value, \
                                       Type=ws['J' + str(i)].value, \
                                       Period=ws['K' + str(i)].value, \
                                       Theme=ws['L' + str(i)].value, \
                                       P1=ws['M' + str(i)].value, \
                                       P2=ws['N' + str(i)].value, \
                                       P3=ws['O' + str(i)].value, \
                                       Logic=ws['P' + str(i)].value, \
                                       Difficulty=ws['Q' + str(i)].value)
                new_paper.add_sub(subject)

                try:
                    (sub_obj, created) = Subject.objects.get_or_create(number=subject.Number)
                    self.copy_sub(sub_obj, subject)
                except Exception as e:
                    sub_obj.delete()
                    print('[Add %d] %s' % (i, repr(e)))
                    continue
                sub_obj.save()
                exam_obj.subjects_set.add(sub_obj)

            print('In paper: %s, total subs: %d' % (new_paper.name, len(new_paper.subjects)))
            exam_obj.save()
            return 0

class SubjectAdmin(admin.ModelAdmin):
    list_display = ('id', 'number', 'type', 'updated_time', 'created_time')
    search_fields = ('id', 'number', 'type')
    fieldsets = (
        (None, {
            'fields': ('number', 'content_txt',
                       'options1', 'options2', 'options3', 'options4', 'options5',
                       'answer_idx', 'period', 'difficult')
        }),
        ('Advanced options', {
            'classes': ('collapse',),
            'fields': ( 'options6', 'options7', 'options8',
                        'content_img', 'content_video',
                        'type', 'theme', 'logic',
                        'point1', 'point2', 'point3'),
        }),
    )
    ordering = ('id',)

'''
class SubjectProxy(Subject):
    class Meta:
        proxy = True
'''

# Register your models here.
admin.site.register(Exam, ExamAdmin)
admin.site.register(Subject, SubjectAdmin)